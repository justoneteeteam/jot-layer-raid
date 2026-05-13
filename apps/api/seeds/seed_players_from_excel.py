import json
import sys
import os
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import SessionLocal, engine, Base
from models.team import Team
from models.player import Player

def seed_players():
    db = SessionLocal()
    try:
        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))), "NFL Database.xlsx")
        wb = openpyxl.load_workbook(excel_path)
        
        # Get all teams from DB to map name to ID
        teams = db.query(Team).all()
        team_map = {t.name.lower(): t.id for t in teams}
        
        total_players = 0
        
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Overall DB' or sheet_name in ('Raven', 'Falcon'):
                continue
                
            # Try to find the corresponding team in DB
            team_id = None
            sheet_lower = sheet_name.lower()
            for t_name_lower, t_id in team_map.items():
                if t_name_lower in sheet_lower or sheet_lower in t_name_lower:
                    team_id = t_id
                    break
            
            # Special cases
            if sheet_name == '49ers':
                team_id = team_map.get('san francisco 49ers')
            elif sheet_name == 'LA Rams':
                team_id = team_map.get('los angeles rams')
                
            if not team_id:
                print(f"Warning: Could not find team ID for sheet '{sheet_name}'. Skipping.")
                continue
                
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            
            # Find the header row (usually row 1)
            header_idx = -1
            for i, row in enumerate(rows):
                if row[0] == 'Name' and row[1] == 'Number':
                    header_idx = i
                    break
                    
            if header_idx == -1:
                print(f"Warning: No valid header found in '{sheet_name}'. Skipping.")
                continue
                
            for row in rows[header_idx + 1:]:
                name = row[0]
                number = row[1]
                type_val = row[2] if len(row) > 2 else "Current"
                group_val = row[3] if len(row) > 3 else "Football"
                
                if not name or number is None:
                    continue
                    
                # Clean up number (could be float like 1.0)
                try:
                    num_val = int(float(number))
                except ValueError:
                    num_val = 0
                    
                player = Player(
                    team_id=team_id,
                    name=str(name).strip(),
                    display_name=str(name).strip().upper(),
                    number=num_val,
                    type=str(type_val) if type_val else "Current",
                    group=str(group_val) if group_val else "Football",
                    is_active=True
                )
                db.add(player)
                total_players += 1
                
        db.commit()
        print(f"✅ Successfully seeded {total_players} players from Excel.")
    except Exception as e:
        db.rollback()
        print(f"❌ Seed failed: {e}")
    finally:
        db.close()

if __name__ == "__main__":
    seed_players()
